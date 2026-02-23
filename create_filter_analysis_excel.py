#!/usr/bin/env python3
"""
Create Filter Analysis Excel File
Replicates the filter analysis spreadsheet structure
"""

import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
import openpyxl

def create_filter_analysis_excel(output_file='Filter_Analysis.xlsx'):
    """Create the filter analysis Excel file with proper formatting"""
    
    # Sample filter data based on the image
    filters = [
        'Status code',
        'current_outstanding_balance_owed_amount_filter',
        'PQ PROXY SCORE',
        'FA PROKY SCORE',
        'FICO_10T',
        'Credit Inquiries',
        'months_since_recent_30plus_delinquency_dereg',
        'PREMIER_V1_1_ALL9210',
        'PREMIER_V1_1_BCC8220',
        'PREMIER_V1_1_ALL0400',
        'UPL5020_plus_BCC5020',
        'term_filter',
        'Modification_status',
        'max_DPD',
        'DPD',
        'state_filter',
        'xpn_ALL6270'
    ]
    
    # Create main dataframe
    df_main = pd.DataFrame({
        'Filter': filters,
        'Exact_Only_Fail_Volume': ['#N/A'] * len(filters),
        'Exact_Only_Fail_%': ['#N/A'] * len(filters),
        'Exact_Only_Fail_Row_Count': ['#VALUE!'] * len(filters),
        'Exact_Only_Fail_Row_%': ['#VALUE!'] * len(filters),
        'Total_Fail_Volume': [126, 0, 980, 483, 26839, 145, 12782, 190, 379, 1154, 22294, 430, 1408, 5320, 3488, 351, 3182],
        'Total_Fail_%_Total': ['0.33%', '0.00%', '44.12%', '16.84%', '69.74%', '0.38%', '33.21%', '0.49%', '0.98%', '3.00%', '57.93%', '1.12%', '3.66%', '13.82%', '9.06%', '0.91%', '8.27%'],
        'Rank_by_Exact_Only_Fail_Volume': ['#N/A'] * len(filters)
    })
    
    # Create Top 10 dataframe
    df_top10 = pd.DataFrame({
        'Rank': range(1, 11),
        'Filter': [''] * 10,
        'Exact_Only_Fail_Volume': ['#N/A'] * 10,
        'Exact_Only_Fail_%': ['#N/A'] * 10
    })
    
    # Create Excel writer
    with pd.ExcelWriter(output_file, engine='openpyxl') as writer:
        df_main.to_excel(writer, sheet_name='Filter Analysis', index=False, startrow=0)
        df_top10.to_excel(writer, sheet_name='Filter Analysis', index=False, startrow=0, startcol=9)
    
    # Load workbook for formatting
    wb = load_workbook(output_file)
    ws = wb['Filter Analysis']
    
    # Define styles
    header_fill = PatternFill(start_color='1F4E78', end_color='1F4E78', fill_type='solid')
    header_font = Font(bold=True, color='FFFFFF', size=11)
    border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )
    
    # Format main table header (row 1, columns A-H)
    for col in range(1, 9):
        cell = ws.cell(row=1, column=col)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal='center', vertical='center')
        cell.border = border
    
    # Format main table data
    for row in range(2, len(filters) + 2):
        for col in range(1, 9):
            cell = ws.cell(row=row, column=col)
            cell.border = border
            if col == 1:  # Filter column
                cell.alignment = Alignment(horizontal='left', vertical='center')
            elif col in [6, 7]:  # Total_Fail_Volume and Total_Fail_%_Total
                cell.alignment = Alignment(horizontal='right', vertical='center')
            else:
                cell.alignment = Alignment(horizontal='center', vertical='center')
    
    # Format Top 10 table header (row 1, columns J-M)
    # Merge cells for "Top 10 Exact Only Fail Filters (17-filter set)" header
    ws.merge_cells('J1:M1')
    header_cell = ws.cell(row=1, column=10)
    header_cell.value = 'Top 10 Exact Only Fail Filters (17-filter set)'
    header_cell.fill = header_fill
    header_cell.font = header_font
    header_cell.alignment = Alignment(horizontal='center', vertical='center')
    header_cell.border = border
    
    # Top 10 table column headers (row 2, columns J-M)
    top10_headers = ['Rank', 'Filter', 'Exact_Only_Fail_Volume', 'Exact_Only_Fail_%']
    for idx, header in enumerate(top10_headers, start=10):
        cell = ws.cell(row=2, column=idx)
        cell.value = header
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal='center', vertical='center')
        cell.border = border
    
    # Format Top 10 table data
    for row in range(3, 13):  # Rows 3-12 for the 10 ranks
        for col in range(10, 14):  # Columns J-M
            cell = ws.cell(row=row, column=col)
            cell.border = border
            if col == 11:  # Filter column - empty with dropdown potential
                cell.alignment = Alignment(horizontal='left', vertical='center')
            else:
                cell.alignment = Alignment(horizontal='center', vertical='center')
    
    # Set column widths
    ws.column_dimensions['A'].width = 45  # Filter
    ws.column_dimensions['B'].width = 22  # Exact_Only_Fail_Volume
    ws.column_dimensions['C'].width = 18  # Exact_Only_Fail_%
    ws.column_dimensions['D'].width = 22  # Exact_Only_Fail_Row_Count
    ws.column_dimensions['E'].width = 20  # Exact_Only_Fail_Row_%
    ws.column_dimensions['F'].width = 18  # Total_Fail_Volume
    ws.column_dimensions['G'].width = 18  # Total_Fail_%_Total
    ws.column_dimensions['H'].width = 30  # Rank_by_Exact_Only_Fail_Volume
    ws.column_dimensions['J'].width = 8   # Rank
    ws.column_dimensions['K'].width = 35  # Filter
    ws.column_dimensions['L'].width = 22  # Exact_Only_Fail_Volume
    ws.column_dimensions['M'].width = 18  # Exact_Only_Fail_%
    
    # Set row height for header
    ws.row_dimensions[1].height = 25
    ws.row_dimensions[2].height = 25
    
    # Freeze panes (freeze first row)
    ws.freeze_panes = 'A2'
    
    wb.save(output_file)
    print(f"Excel file created: {output_file}")
    return output_file


if __name__ == "__main__":
    create_filter_analysis_excel()
